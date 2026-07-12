import os
import openpyxl

_candidates = ['/mnt/c/Users/yn441/Downloads/技術経歴書.xlsx', os.path.expanduser('~/Downloads/技術経歴書.xlsx')]
path = next((p for p in _candidates if os.path.exists(p)), _candidates[0])

wb = openpyxl.load_workbook(path, data_only=True)
for sh in wb.sheetnames:
    ws = wb[sh]
    print(f"=== {sh} ===")
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            print(row)
